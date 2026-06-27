# users/services.py

from django.db import transaction
from django.contrib.auth.hashers import make_password
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from users.models import Employee, Role, Subdepartment
import logging

logger = logging.getLogger(__name__)

# ستون‌های مورد نیاز در فایل اکسل
REQUIRED_COLUMNS = [
    "national_id",
]
# REQUIRED_COLUMNS = [
#     'national_id', 'personnel_code'
# ]
OPTIONAL_COLUMNS = [
    "first_name",
    "last_name",
    "phone_number",
    "subdepartments_list",
    "roles_list",
    "password",
    "personnel_code",
    "center_of_charge",
    "Position_Job_history",
    "Place_of_service_recruitment_order",
    "is_active",
]


def parse_id_list(id_str):
    """
    تبدیل رشته آیدی‌ها (مثل '1,2;3' یا ' 4 , 5 ') به لیست اعداد صحیح
    """
    if not id_str:
        return []
    # جایگزینی نقطه‌ویرگول با کاما و تقسیم
    parts = [p.strip() for p in id_str.replace(";", ",").split(",") if p.strip()]
    ids = []
    for p in parts:
        try:
            ids.append(int(p))
        except ValueError:
            return None, f"آیدی نامعتبر: '{p}' (باید عدد صحیح باشد)"
    return ids, None


def import_employees_from_excel(excel_file):
    """
    وارد کردن کارمندان از اکسل با استفاده از آیدی برای زیربخش‌ها و نقش‌ها
    """
    imported_count = 0
    skipped_count = 0
    errors = []

    # بررسی حجم فایل
    if excel_file.size < 500:
        return {
            "success": 0,
            "skipped": 0,
            "errors": [f"فایل بیش از حد کوچک است. (حجم: {excel_file.size} بایت)"],
        }

    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except InvalidFileException:
        return {"success": 0, "skipped": 0, "errors": ["فایل XLSX معتبر نیست."]}
    except Exception as e:
        return {"success": 0, "skipped": 0, "errors": [f"خطا در باز کردن فایل: {e}"]}

    if sheet.max_row < 2:
        return {
            "success": 0,
            "skipped": 0,
            "errors": ["فایل اکسل خالی است یا فقط سرستون دارد."],
        }

    # خواندن سرستون‌ها
    headers = [cell.value.strip() if cell.value else "" for cell in sheet[1]]
    col_indices = {}

    # MARK: todo : update dose not need these REQUIRED COLUMNS

    for col in REQUIRED_COLUMNS:
        if col not in headers:
            errors.append(f"ستون مورد نیاز '{col}' در فایل اکسل پیدا نشد.")
            return {"success": 0, "skipped": 0, "errors": errors}
        col_indices[col] = headers.index(col)

    for col in OPTIONAL_COLUMNS:
        if col in headers:
            col_indices[col] = headers.index(col)

    # پردازش سطرها
    with transaction.atomic():
        exist_count = 0
        new_count = 0
        # numb = 0
        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            data = {}
            row_errors = []

            # استخراج داده‌ها
            for col_name, col_idx in col_indices.items():
                value = row[col_idx].value
                data[col_name] = str(value).strip() if value is not None else ""

            # اعتبارسنجی فیلدهای ضروری
            national_id = data.get("national_id", "").strip()
            if not national_id or national_id == "":
                row_errors.append("کد ملی نمی‌تواند خالی باشد.")
                continue

            employee = Employee.objects.filter(national_id=data["national_id"]).first()

            

            if employee or employee != None:
                # exist_count += 1
                # continue

                subdept_objs = []
                if "subdepartments_list" in data and data["subdepartments_list"]:
                    subdept_ids, err = parse_id_list(data["subdepartments_list"])
                    if err:
                        row_errors.append(err)
                    else:
                        for sid in subdept_ids:
                            try:
                                subdept = Subdepartment.objects.get(id=sid)
                                subdept_objs.append(subdept)
                            except Subdepartment.DoesNotExist:
                                row_errors.append(f"زیربخش با آیدی {sid} وجود ندارد.")
                            except ValueError:
                                row_errors.append(f"آیدی زیربخش {sid} نامعتبر است.")
                else:
                    subdept_objs = []

                role_objs = []
                if "roles_list" in data and data["roles_list"]:
                    role_ids, err = parse_id_list(data["roles_list"])
                    if err:
                        row_errors.append(err)
                    else:
                        for rid in role_ids:
                            try:
                                role = Role.objects.get(id=rid)
                                role_objs.append(role)
                            except Role.DoesNotExist:
                                row_errors.append(f"نقش با آیدی {rid} وجود ندارد.")
                            except ValueError:
                                row_errors.append(f"آیدی نقش {rid} نامعتبر است.")
                else:
                    role_objs = []  # خالی بودن مجاز است

                if row_errors:
                    errors.extend([f"ردیف {row_index}: {err}" for err in row_errors])
                    skipped_count += 1
                    continue

                data["personnel_code"] = data.get("personnel_code", "")

                first_name = data.get("first_name", "").strip()
                if first_name or not str(
                    first_name
                ).strip().lower() in ["none", "nan", "null", ""]:
                    employee.first_name = first_name

                last_name = data.get("last_name", "").strip()
                if last_name or not str(last_name).strip().lower() in [
                    "none",
                    "nan",
                    "null",
                    "",
                ]:
                    employee.last_name = last_name

                phone_number = data.get("phone_number", "").strip()
                if phone_number or not str(
                    phone_number
                ).strip().lower() in ["none", "nan", "null", ""]:
                    employee.phone_number = phone_number

                personnel_code = data.get("personnel_code", "").strip()
                if personnel_code or not str(
                    personnel_code
                ).strip().lower() in ["none", "nan", "null", ""]:
                    employee.personnel_code = personnel_code

                password = data.get("password", "").strip()
                if password or not str(password).strip().lower() in [
                    "none",
                    "nan",
                    "null",
                    "",
                ]:
                    # هش کردن رمز عبور
                    hashed_password = make_password(password)
                    employee.password = hashed_password

                center_of_charge = data.get("center_of_charge", "").strip()
                if center_of_charge or not str(
                    center_of_charge
                ).strip().lower() in ["none", "nan", "null", ""]:
                    employee.center_of_charge = center_of_charge

                Position_Job_history = data.get("Position_Job_history", "").strip()
                if Position_Job_history or not str(
                    Position_Job_history
                ).strip().lower() in ["none", "nan", "null", ""]:
                    employee.Position_Job_history = Position_Job_history
                
                Place_of_service_recruitment_order = data.get("Place_of_service_recruitment_order", "").strip()
                if Place_of_service_recruitment_order or not str(
                    Place_of_service_recruitment_order
                ).strip().lower() in ["none", "nan", "null", ""]:
                    employee.Place_of_service_recruitment_order = Place_of_service_recruitment_order

                is_active = data.get("is_active", "").strip()
                if is_active or not str(
                    is_active
                ).strip().lower() in ["none", "nan", "null", ""]:
                    if is_active == "فعال":
                        employee.is_active = True
                    elif is_active == "غیر فعال":
                        employee.is_active = False

                employee.save()

                subdepartments_list = data.get("subdepartments_list", "").strip()
                if subdepartments_list or not subdepartments_list in [
                    "none",
                    "nan",
                    "null",
                    "",
                ]:
                    employee.assigned_subdepartments.set(subdept_objs)

                roles_list = data.get("roles_list", "").strip()
                if roles_list or not roles_list in [
                    "none",
                    "nan",
                    "null",
                    "",
                ]:
                    employee.roles.set(role_objs)

                imported_count += 1

            else:

                # is_active = data.get("is_active", "").strip()
                # if is_active or not str(
                #     is_active
                # ).strip().lower() in ["none", "nan", "null", ""]:
                #     if is_active != "غیر فعال":
                #         new_count += 1
                # continue

                is_active = data.get("is_active", "").strip()
                if is_active or not str(
                    is_active
                ).strip().lower() in ["none", "nan", "null", ""]:
                    if is_active != "غیر فعال":
                        skipped_count += 1
                        continue

                if not data["first_name"] or not data["last_name"]:
                    row_errors.append("نام یا نام خانوادگی نمی‌تواند خالی باشد.")


                password = data.get("password", "").strip()
                if not password or str(password).strip().lower() in [
                    "none",
                    "nan",
                    "null",
                    "",
                ]:
                    password = "123456"
                    # row_errors.append("رمز عبور نمی‌تواند خالی یا نامعتبر باشد.")
                    # errors.append(
                    #     f"ردیف {row_index}: رمز عبور نمی‌تواند خالی یا نامعتبر باشد."
                    # )
                    # skipped_count += 1
                    # continue  # بدون این continue ورکر می‌میرد!

                if row_errors:
                    errors.append(f"ردیف {row_index}: {' | '.join(row_errors)}")
                    skipped_count += 1
                    continue

                subdept_objs = []
                if "subdepartments_list" in data and data["subdepartments_list"]:
                    subdept_ids, err = parse_id_list(data["subdepartments_list"])
                    if err:
                        row_errors.append(err)
                    else:
                        for sid in subdept_ids:
                            try:
                                subdept = Subdepartment.objects.get(id=sid)
                                subdept_objs.append(subdept)
                            except Subdepartment.DoesNotExist:
                                row_errors.append(f"زیربخش با آیدی {sid} وجود ندارد.")
                            except ValueError:
                                row_errors.append(f"آیدی زیربخش {sid} نامعتبر است.")
                else:
                    subdept_objs = []  # خالی بودن مجاز است

                # --- پردازش roles_list (اختیاری) ---
                role_objs = []
                if "roles_list" in data and data["roles_list"]:
                    role_ids, err = parse_id_list(data["roles_list"])
                    if err:
                        row_errors.append(err)
                    else:
                        for rid in role_ids:
                            try:
                                role = Role.objects.get(id=rid)
                                role_objs.append(role)
                            except Role.DoesNotExist:
                                row_errors.append(f"نقش با آیدی {rid} وجود ندارد.")
                            except ValueError:
                                row_errors.append(f"آیدی نقش {rid} نامعتبر است.")
                else:
                    role_objs = []  # خالی بودن مجاز است

                if row_errors:
                    errors.extend([f"ردیف {row_index}: {err}" for err in row_errors])
                    skipped_count += 1
                    continue

                data["personnel_code"] = data.get("personnel_code", "")
                # هش کردن رمز عبور
                hashed_password = make_password(password)

                try:
                    # ایجاد یا به‌روزرسانی بر اساس national_id
                    employee, created = Employee.objects.get_or_create(
                        national_id=data["national_id"],
                        defaults={
                            "first_name": data["first_name"],
                            "last_name": data["last_name"],
                            "phone_number": data["phone_number"],
                            "personnel_code": data["personnel_code"],
                            "password": hashed_password,
                            "is_active": True,
                            "is_first_login": True,
                        },
                    )

                    # به‌روزرسانی در صورت وجود
                    if not created:
                        employee.first_name = data["first_name"]
                        employee.last_name = data["last_name"]
                        employee.phone_number = data["phone_number"]
                        if data[
                            "personnel_code"
                        ]:  # فقط اگر مقدار جدید داشته باشه، به‌روزرسانی کن
                            employee.personnel_code = data["personnel_code"]
                        employee.password = hashed_password
                        employee.is_first_login = True
                        employee.save()

                    # همگام‌سازی ManyToMany با set()
                    employee.assigned_subdepartments.set(subdept_objs)
                    employee.roles.set(role_objs)

                    imported_count += 1

                except Exception as e:
                    errors.append(f"ردیف {row_index}: خطا در ذخیره کارمند - {e}")
                    skipped_count += 1

    return {"success": imported_count, "skipped": skipped_count, "errors": errors}
    # print(f"new: {new_count}, exist: {exist_count}")
    # return {"success": new_count, "skipped": exist_count, "errors": ""}
