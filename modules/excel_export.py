from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import jdatetime
from datetime import datetime
from django.http import HttpResponseBadRequest


def export_to_excel(
    columns,
    filename="گزارش.xlsx",
    report_title="گزارش سیستم",
    restaurant_stats=None,
    summary_row=None,
):
    print(columns)

    if not columns:
        return HttpResponseBadRequest("داده‌ای برای اکسپورت وجود ندارد.")

    row_count = len(columns[0][1])
    for title, values in columns:
        if len(values) != row_count:
            raise ValueError("تعداد ردیف‌ها در همه ستون‌ها باید یکسان باشد.")

    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش"

    header_font = Font(name="B Nazanin", size=12, bold=True, color="FFFFFF")
    cell_font = Font(name="B Nazanin", size=11)
    summary_font = Font(name="B Nazanin", size=11, bold=True, color="FF0000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill("solid", fgColor="1f4e79")

    row = 1

    # عنوان گزارش
    ws.cell(row=row, column=1, value=report_title).font = Font(
        name="B Nazanin", size=16, bold=True
    )
    ws.cell(row=row, column=1).alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    row += 2

    # تاریخ تولید
    jalali_date = jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M")
    ws.cell(row=row, column=1, value=f"تاریخ تولید گزارش: {jalali_date}").font = Font(
        name="B Nazanin", size=10, italic=True
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    row += 2

    # سرستون‌ها
    for col_num, (header_text, _) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    row += 1

    # داده‌ها
    for i in range(row_count):
        for col_num, (_, values) in enumerate(columns, 1):
            value = values[i] if i < len(values) else ""
            if value is None:
                value = ""
            if isinstance(value, datetime):
                value = jdatetime.date.fromgregorian(date=value.date()).strftime(
                    "%Y/%m/%d"
                )

            cell = ws.cell(row=row, column=col_num, value=str(value))
            cell.font = cell_font
            cell.border = border
            cell.alignment = center if "تاریخ" in columns[col_num - 1][0] else right
        row += 1

    # اضافه کردن ردیف خلاصه برای رستوران‌ها
    if summary_row and restaurant_stats:
        # یک ردیف خالی
        row += 1

        # عنوان خلاصه
        summary_title_cell = ws.cell(
            row=row, column=1, value="خلاصه تعداد کل سفارشات هر رستوران:"
        )
        summary_title_cell.font = Font(name="B Nazanin", size=12, bold=True)
        summary_title_cell.alignment = Alignment(horizontal="right")

        # اگر ستون‌های رستوران داریم، عنوان خلاصه را Merge می‌کنیم
        if len(columns) > 10:  # اگر ستون‌های رستوران وجود دارند
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

        row += 1

        # داده‌های خلاصه هر رستوران
        # شروع از ستون 9 (ستون I) زیرا 8 ستون اول داریم
        for i, restaurant in enumerate(restaurant_stats.values(), start=10):
            if i <= len(columns):  # مطمئن شویم ستون وجود دارد
                col_letter = get_column_letter(i)
                cell = ws[f"{col_letter}{row}"]
                cell.value = restaurant["total_orders"]
                cell.font = summary_font
                cell.alignment = center
                cell.border = border

    for r in range(1, row):
        ws.row_dimensions[r].height = 28

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    safe_filename = filename if filename.endswith(".xlsx") else f"{filename}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{safe_filename}"'
    wb.save(response)
    return response
